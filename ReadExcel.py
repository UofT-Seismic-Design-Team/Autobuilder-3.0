import os
import win32com.client
import random
from openpyxl import *
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import string

class Tower:
    def __init__(self, number=1, x_width=0, y_width=0, panels={}, members={}):
        self.number = number
        self.x_width = x_width
        self.y_width = y_width
        self.panels = panels
        self.members = members

class BracingScheme:
    def __init__(self, number=1, members=[], mass_nodes=[]):
        self.number = number
        self.members = members
        self.mass_nodes = mass_nodes

class FloorPlan:
    def __init__(self, number=1, members=[], mass_nodes=[], scaling_x = 0, scaling_y = 0, area = 0):
        self.number = number
        self.members = members
        self.mass_nodes = mass_nodes
        self.scaling_x = scaling_x
        self.scaling_y = scaling_y
        self.area = area

class Member:
    def __init__(self, start_node=[], end_node=[], sec_prop=1):
        self.start_node = start_node
        self.end_node = end_node
        self.sec_prop = sec_prop

class Panel:
    def __init__(self, num, point1, point2, point3, point4):
        self.num = num
        self.point1 = point1 # Lower left
        self.point2 = point2 # Upper left
        self.point3 = point3 # Upper right
        self.point4 = point4 # Lower right


##########get excel index############
def get_excel_indices(wb, index_headings_col, index_values_col, index_start_row):
    excel_index = {}
    ws = wb['Index']
    current_row = index_start_row
    while ws[index_headings_col + str(current_row)].value is not None:
        index_heading = ws[index_headings_col + str(current_row)].value
        index_value = ws[index_values_col + str(current_row)].value
        #enter the new entry into the index
        excel_index[index_heading] = index_value
        current_row = current_row + 1
    return excel_index

##########read excel tabs############
def get_properties(wb,excel_index,parameter):
    headings_start_col = excel_index['Section or material properties col']
    values_start_col = excel_index['Section or material values col']
    start_row = excel_index['Properties start row']
    if parameter == 'Material':
        ws = wb['Materials']
    elif parameter == 'Section':
        ws = wb['Section Properties']
    else:
        print('Input should be either "Material" or "Section"')
    #parameter = 'unknown'
    #if ws['A1'].value == 'Section #':
    #    parameter = 'Section'
    #else:
    #    parameter = 'Material'
    parameter_type = {};
    current_property_col = headings_start_col;
    current_value_col = values_start_col;
    i = 1 
    while ws[current_property_col+str(1)].value is not None:
        current_row = start_row
        parameter_type[parameter+' '+str(i)]={}
        while ws[current_property_col + str(current_row)].value is not None:
            properties = {}
            properties_heading = ws[current_property_col + str(current_row)].value
            properties_value = ws[current_value_col + str(current_row)].value
            #enter the new entry into the index
            parameter_type[parameter+' '+str(i)][properties_heading] = properties_value
            current_row = current_row + 1
        i += 1
        current_property_col = get_column_letter(column_index_from_string(current_property_col)+3)
        current_value_col = get_column_letter(column_index_from_string(current_value_col)+3)
    return parameter_type

#Outputs a list containing tower objects representing each tower to be built
def read_input_table(wb,excel_index):
    #Set worksheet to the input table sheet
    ws_input = wb['Input Table']
    input_table_offset = excel_index['Input table offset']
    total_towers = excel_index['Total number of towers']
    cur_tower_row = 1
    all_towers = []
    panel_num_col = 'A'
    panel_bracing_col = 'B'
    member_name_col = 'C'
    member_prop_col = 'D'

    while ws_input[panel_num_col + str(cur_tower_row)].value is not None:
        # Read tower number
        cur_tower_num = int(ws_input[panel_bracing_col + str(cur_tower_row)].value)
        # Read tower x width
        x_width = float(ws_input[panel_bracing_col + str(cur_tower_row + 1)].value)
        # Read tower y width
        y_width = ws_input[panel_bracing_col + str(cur_tower_row + 2)].value
        # Read panels
        panels = {}
        cur_panel_row = cur_tower_row + input_table_offset
        while ws_input[panel_num_col + str(cur_panel_row)].value is not None:
            panel_num = ws_input[panel_num_col + str(cur_panel_row)].value
            panel_bracing = ws_input[panel_bracing_col + str(cur_panel_row)].value
            panels[panel_num] = panel_bracing
            cur_panel_row += 1
        # Read members
        members = {}
        cur_member_row = cur_tower_row + input_table_offset
        while ws_input[member_name_col + str(cur_member_row)].value is not None:
            member_name = ws_input[member_name_col + str(cur_member_row)].value
            sec_prop = ws_input[member_prop_col + str(cur_member_row)].value
            members[member_name] = sec_prop
            cur_member_row += 1
        # Increment
        cur_tower = Tower(number=cur_tower_num, x_width=x_width, y_width=y_width, panels=panels, members=members)
        all_towers.append(cur_tower)
        cur_tower_row = max(cur_panel_row, cur_member_row) + 1
        if ws_input[panel_num_col + str(cur_tower_row)].value is None and ws_input[get_column_letter(column_index_from_string(panel_num_col)+5) + str(1)] is not None:
            panel_num_col = get_column_letter(column_index_from_string(panel_num_col)+5)
            panel_bracing_col = get_column_letter(column_index_from_string(panel_bracing_col)+5)
            member_name_col = get_column_letter(column_index_from_string(member_name_col)+5)
            member_prop_col = get_column_letter(column_index_from_string(member_prop_col)+5)
            cur_tower_row = 1
    print('Read ' + str(len(all_towers)) + ' towers')
    return all_towers

def get_panels(wb, excel_index):
    cur_panel_col = 'A'
    ws = wb['Panels']
    panels = []
    panel_num = 1
    while ws[cur_panel_col + '1'].value is not None:
        xCol = get_column_letter(column_index_from_string(cur_panel_col) + 1)
        yCol = get_column_letter(column_index_from_string(cur_panel_col) + 2)
        zCol = get_column_letter(column_index_from_string(cur_panel_col) + 3)
        x1 = ws[xCol + '4'].value
        x2 = ws[xCol + '5'].value
        x3 = ws[xCol + '6'].value
        x4 = ws[xCol + '7'].value
        y1 = ws[yCol + '4'].value
        y2 = ws[yCol + '5'].value
        y3 = ws[yCol + '6'].value
        y4 = ws[yCol + '7'].value
        z1 = ws[zCol + '4'].value
        z2 = ws[zCol + '5'].value
        z3 = ws[zCol + '6'].value
        z4 = ws[zCol + '7'].value
        panels.append(Panel(num=panel_num, point1=[x1,y1,z1], point2=[x2,y2,z2], point3=[x3,y3,z3], point4=[x4,y4,z4]))
        panel_num += 1
        cur_panel_col = get_column_letter(column_index_from_string(cur_panel_col) + 5)
    print('Read ' + str(len(panels)) + ' panels')
    return panels

def get_node_info(wb,excel_index,node_num_col,parameter):
    if parameter == 'Floor Bracing':
        ws = wb['Floor Bracing']
    elif parameter == 'Bracing':
        ws = wb['Bracing']
    elif parameter == 'Space Bracing':
        ws = wb['Space Bracing']
    elif parameter == 'Floor Plans':
        ws = wb['Floor Plans']
    horiz_col = get_column_letter(column_index_from_string(node_num_col)+1)
    vert_col = get_column_letter(column_index_from_string(node_num_col)+2)
    mass_col = get_column_letter(column_index_from_string(node_num_col)+3)
    start_row = excel_index['Properties start row']
    nodes = []
    mass_nodes = []
    current_row = start_row
    while ws[node_num_col + str(current_row)].value is not None:
        horiz = ws[horiz_col + str(current_row)].value
        vert = ws[vert_col + str(current_row)].value
        mass_at_node = ws[mass_col + str(current_row)].value
        if mass_at_node ==1:
            mass_nodes.append([horiz,vert])
        #enter the new entry into the list
        nodes.append([horiz, vert])
        current_row = current_row + 1
    return nodes, mass_nodes

def get_bracing(wb,excel_index,parameter):
    headings_col = excel_index['Bracing start col']
    section_col = excel_index['Bracing section col']
    start_node_col = excel_index['Bracing start node col']
    end_node_col = excel_index['Bracing end node col']
    start_row = excel_index['Properties start row']
    if parameter == 'Floor Bracing':
        ws = wb['Floor Bracing']
    elif parameter == 'Bracing':
        ws = wb['Bracing']
    elif parameter == 'Space Bracing':
        ws = wb['Space Bracing']
    else:
        print('Input should be either "Floor Bracing", "Space Bracing" or "Bracing"')
    all_bracing = []
    current_headings_col = headings_col
    current_section_col = section_col
    current_start_node_col = start_node_col
    current_end_node_col = end_node_col
    i = 1
    while ws[current_headings_col+str(4)].value is not None:
        nodes, mass_nodes = get_node_info(wb, excel_index, current_headings_col, parameter)
        current_row = start_row
        j = 1
        cur_members = []
        while ws[current_start_node_col + str(current_row)].value is not None:
            section = ws[current_section_col + str(current_row)].value
            start_node_num = ws[current_start_node_col + str(current_row)].value
            end_node_num = ws[current_end_node_col + str(current_row)].value
            start_node = nodes[start_node_num-1]
            end_node = nodes[end_node_num-1]
            cur_members.append(Member(start_node, end_node, section))
            current_row = current_row + 1
            j += 1
        all_bracing.append(BracingScheme(number=i, members=cur_members, mass_nodes=mass_nodes))
        i += 1
        current_headings_col = get_column_letter(column_index_from_string(current_headings_col)+9)
        current_section_col = get_column_letter(column_index_from_string(current_section_col)+9)
        current_start_node_col = get_column_letter(column_index_from_string(current_start_node_col)+9)
        current_end_node_col = get_column_letter(column_index_from_string(current_end_node_col)+9)
    print('Read ' + str(len(all_bracing)) + ' bracing schemes')
    return all_bracing

def get_floor_plans(wb,excel_index):
    headings_col = excel_index['Floor plan start col']
    section_col = excel_index['Floor plan section col']
    start_node_col = excel_index['Floor plan start node col']
    end_node_col = excel_index['Floor plan end node col']
    start_row = excel_index['Properties start row']
    ws = wb['Floor Plans']
    all_plans = []
    current_headings_col = headings_col
    current_section_col = section_col
    current_start_node_col = start_node_col
    current_end_node_col = end_node_col
    i = 1
    while ws[current_headings_col + str(4)].value is not None:
        [nodes, mass_nodes] = get_node_info(wb, excel_index, current_headings_col, 'Floor Plans')
        current_row = start_row
        cur_members = []
        max_node_x = 0
        max_node_y = 0
        min_node_x = 0
        min_node_y = 0
        while ws[current_start_node_col + str(current_row)].value is not None:
            section = ws[current_section_col + str(current_row)].value
            start_node_num = ws[current_start_node_col + str(current_row)].value
            end_node_num = ws[current_end_node_col + str(current_row)].value
            start_node = nodes[start_node_num - 1]
            end_node = nodes[end_node_num - 1]
            cur_members.append(Member(start_node, end_node, section))
            #find scaling factor in x and y, find area
            cur_nodes = []
            cur_nodes.append(start_node)
            cur_nodes.append(end_node)
            for node in cur_nodes:
                if max_node_x < start_node[0]:
                    max_node_x = node[0]
                if max_node_y < node[1]:
                    max_node_y = node[1]
                if min_node_x > node[0]:
                    min_node_x = node[0]
                if min_node_y > node[1]:
                    min_node_y = node[1]
            current_row = current_row + 1

        scaling_x = max_node_x - min_node_x
        scaling_y = max_node_y - min_node_y
        #area = scaling_x*scaling_y
        
        all_plans.append(FloorPlan(number=i, members=cur_members, mass_nodes=mass_nodes, scaling_x = scaling_x, scaling_y = scaling_y))
        i += 1
        current_headings_col = get_column_letter(column_index_from_string(current_headings_col) + 9)
        current_section_col = get_column_letter(column_index_from_string(current_section_col) + 9)
        current_start_node_col = get_column_letter(column_index_from_string(current_start_node_col) + 9)
        current_end_node_col = get_column_letter(column_index_from_string(current_end_node_col) + 9)
    return all_plans


#TESTING
#wb = load_workbook('SetupAB.xlsm')
#ExcelIndex = get_excel_indices(wb, 'A', 'B', 2)

#InputTable = ExcelIndex['Input table sheet']
#FloorPlan = ExcelIndex['Floor plans sheet']
#SectionProperties = ExcelIndex['Section properties sheet']
#Bracing = ExcelIndex['Bracing sheet'
#FloorBracing = ExcelIndex['Floor bracing sheet']
#Materials = ExcelIndex['Materials sheet']
#InputTableOffset = ExcelIndex['Input table offset']
#PropertiesStartRow = ExcelIndex['Properties start row']

#sectionproperties = get_properties(wb,excelindex,'section')
#materials = get_properties(wb,excelindex,'material')
#bracingschemes = get_bracing(wb,excelindex,'bracing')


#AllFloorPlans = get_floor_plans(wb,ExcelIndex)

#FloorBracing = get_floor_or_bracing(wb,ExcelIndex,'Floor Bracing')

#for keys,values in ExcelIndex.items():
#    print(keys)
#    print(values)

#for keys,values in SectionProperties.items():
#    print(keys)
#    print(values)

#for keys,values in Materials.items():
#    print(keys)
#    print(values)

#for keys,values in Bracing.items():
#   print(keys) 
#   print(values)


#for FloorPlan in AllFloorPlans:
#    print('number ' + str(FloorPlan.number))
#    for member in FloorPlan.members:
#        print(member.start_node)
#        print(member.end_node)
#        print(member.sec_prop)
#    print(FloorPlan.mass_nodes)
#    print(FloorPlan.scaling_x)
#    print(FloorPlan.scaling_y)
#    print(FloorPlan.area)

#for Scheme in BracingSchemes:
    #print('number ' + str(Scheme.number))
    #for member in Scheme.members:        
        #print(member.start_node)
        #print(member.end_node)
        #print(member.sec_prop)


#AllTowers = read_input_table(wb, ExcelIndex)
#for tower in AllTowers:
#    print(tower.number)
#    print(tower.member_mat)
#    print(tower.rod_mat)
#    print(tower.floor_plans)
#    print(tower.floor_heights)
#    print(tower.col_props)
#    print(tower.bracing_types)
#    print(tower.floor_masses)
#    print(tower.floor_bracing_types)
