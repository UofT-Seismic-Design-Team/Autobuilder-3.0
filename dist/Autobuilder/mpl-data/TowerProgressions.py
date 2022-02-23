import xlrd
from openpyxl import Workbook

class Variable:
    def __init__(self, variable_value, variable_type, constraints = [''], autobuilder_values = [''], autobuilder_variables = ['']):
        self.variable_value = variable_value
        self.variable_type = variable_type
        self.constraints = constraints
        self.autobuilder_values = autobuilder_values
        self.autobuilder_variables = autobuilder_variables
        
    def __repr__(self):
        return self.variable_value
    
class Progression:
    def __init__(self, variable_order, tower_numbers, period, cr, cost, dev_period = None, mean_cr = None, mean_cost = None, normal_period = None, normal_cr = None, normal_cost = None):
        self.variable_order = variable_order
        self.tower_numbers = tower_numbers
        self.period = period
        self.cr = cr
        self.cost = cost
        self.dev_period = dev_period
        self.mean_cr = mean_cr
        self.mean_cost = mean_cost
        self.normal_period = normal_period
        self.normal_cr = normal_cr
        self.normal_cost = normal_cost        
        
    def __repr__(self):
        return str(self.variable_order)
    
def find_tower(results, variable_dict, header_dict):
            for row in results:
                match = True
                for var in variable_dict:
                    if not (row[header_dict[var]] == variable_dict[var]):
                        match = False
                if match == True:
                    '''
                    number = (row[header_dict['Tower #']])
                    period = float(row[header_dict['Period']])
                    error = float(row[header_dict['Error']])
                    cost = float(row[header_dict['Cost']])
                    '''
                    number = (row[header_dict['towerNum']])
                    period = float(row[header_dict['period']])
                    error = float(row[header_dict['avgEcc']])
                    cost = float(row[header_dict['totalWeight']])                    
                    return [number, period, error, cost]
            return [None, 0.15, 1.0, 1000.0]

def std_deviation_period(progression):
    period = progression.period
    if len(period) > 1:
        length = len(period)
        delta_period = []
        for i in range(length-1):
            delta_period.append((period[i+1]-period[i]))
        #calculating mean
        sigma1 = 0
        for number in delta_period:
            sigma1 += number
        mean = float(sigma1/length)
        #calculating std deviation
        sigma2 = 0
        for number in delta_period:
            sigma2 += float((number - mean)**2)
        std_deviation = float((sigma2/(length-1))**0.5)
        return std_deviation
    else:
        return 0
    
def mean_cost(progression):
    cost = progression.cost
    length = len(cost)
    #calculating mean
    sigma1 = 0
    for number in cost:
        sigma1 += number
    mean = float(sigma1/length)
    return mean

def mean_cr(progression):
    cr = progression.cr
    length = len(cr)
    #calculating mean
    sigma1 = 0
    for number in cr:
        sigma1 += number
    mean = float(sigma1/length)
    return mean

class TowerProgressions():

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        #file locations for input and output spreadsheets
        input_sheet = 'Progression Input Spreadsheet.xlsx'
        #input_sheet_constants = 'Input Spreadsheet 2.xlsx'
        autobuilder_sheet = 'Progression Autobuilder Results.xlsx'
        output_sheet = "Progression Output Spreadsheet.xlsx"
        
        #normailizing variables
        normal_period_A = 0
        normal_period_B = 0.015
        normal_cr_A = 0
        normal_cr_B = 2
        normal_cost_A = 0
        normal_cost_B = 2000
        
        #reading input
        workbook = xlrd.open_workbook(input_sheet)
        sheet = workbook.sheet_by_index(0)
        
        number_of_varibles = sheet.nrows - 1
        
        #MAKE LEAST TO GREATEST STIFFNESS LIST OF EACH VAR#
        
        list_of_varibleValues = []
        list_of_varibleTypes = []
        list_of_constriants = []
        list_of_autoValues = []
        list_of_autoVaribles = []
        
            
        for i in range(number_of_varibles):
            list_of_varibleValues.append(sheet.cell_value(i+1,1))
            list_of_varibleTypes.append(sheet.cell_value(i+1,2))
        
            temp = sheet.cell_value(i+1,3)
            temp_2 = temp.split(",")
            list_of_constriants.append(temp_2)
            
            temp1 = sheet.cell_value(i+1,4)
            temp1_2 = temp1.split(",")
            list_of_autoValues.append(temp1_2)
        
            temp2 = sheet.cell_value(i+1,5)
            temp2_2 = temp2.split(",")
            list_of_autoVaribles.append(temp2_2)
        
        
        variable_list = []
        
        for i in range(number_of_varibles):
            globals()["object"+str(i+1)] = Variable(list_of_varibleValues[i],list_of_varibleTypes[i],list_of_constriants[i],list_of_autoValues[i],list_of_autoVaribles[i])
            variable_list.append(Variable(list_of_varibleValues[i],list_of_varibleTypes[i],list_of_constriants[i],list_of_autoValues[i],list_of_autoVaribles[i]))
        
        print('generating progressions...')
        
        #create base tower
        base_tower = []
        for variable in variable_list:
            if variable.constraints == ['']:
                base_tower.append(variable)
        
        #create list of different variable types
        variable_types = []
        for variable in variable_list:
            if variable.variable_type not in variable_types:
                variable_types.append(variable.variable_type)
        
        #find number of steps in each progression
        total_steps = len(variable_list) - len(variable_types)
        
        #create list of of progression lists
        prev_step = [base_tower]
        for n in range(total_steps):
            current_step = []
            for i in prev_step:
                for var in variable_list:
                    add_step = True
                    for j in i:
                        if var == j:
                            add_step = False
                    pass_constraints = []
                    for constraint in var.constraints:
                        current_pass = False
                        for k in i:
                            if constraint == k.variable_value:
                                current_pass = True
                        pass_constraints.append(current_pass)
                    if (add_step == True) and (False not in pass_constraints):
                        current_progression = []
                        for l in i:
                            current_progression.append(l)
                        current_progression.append(var)
                        current_step.append(current_progression)
            prev_step = current_step
            print('current progressions: ' + str(len(prev_step)))
        
        #create list of progression objects 
        progression_list = []
        for i in prev_step:
            temp_list = i[len(variable_types):]
            temp_prog = Progression(temp_list, [], [], [], [])
            progression_list.append(temp_prog)
        
        print(str(len(progression_list)) + ' progressions generated')
        
        #read autobuilder results
        print('reading autobuilder results...')
        wb2 = xlrd.open_workbook(autobuilder_sheet)
        ws2 = wb2.sheet_by_index(0)
        rows = ws2.nrows
        cols = ws2.ncols
        autobuilder_headers = []
        for column in range(cols):
            autobuilder_headers.append(ws2.cell(0,column).value)
        autobuilder_results = []
        for row in range(1, rows):
            autobuilder_results.append([])
            for column in range(cols):
                if type(ws2.cell(row,column).value) == float:
                    cur_cell = (ws2.cell(row,column).value)
                    if cur_cell % 1 == 0:
                        int_cell = int(cur_cell)
                        autobuilder_results[row-1].append(str(int_cell))
                    else:
                        autobuilder_results[row-1].append(str(cur_cell))
                else:
                    autobuilder_results[row-1].append(ws2.cell(row,column).value)
                
        #write autobuilder headers to dictionary
        print('matching towers to autobuilder...')
        header_count = 0
        header_dict = {}
        for header in autobuilder_headers:
            header_dict.update({header : header_count})
            header_count += 1
        
        #write base tower values to dictionary and match to autobuilder
        tower_dict = {}
        for var in base_tower:
            for auto_var in range(len(var.autobuilder_variables)):
                tower_dict.update({var.autobuilder_variables[auto_var] : var.autobuilder_values[auto_var]})
        base_values = find_tower(autobuilder_results, tower_dict, header_dict)
        
        
        #loop through all progressions and find matching autobuilder results
        for progression in progression_list:
            progression.tower_numbers.append(base_values[0])
            progression.period.append(base_values[1])
            progression.cr.append(base_values[2])
            progression.cost.append(base_values[3])
            current_tower = {}
            for i in tower_dict:
                current_tower.update({i : tower_dict[i]})
            for tower in progression.variable_order:
                for auto_var in range(len(tower.autobuilder_variables)):
                    current_tower[tower.autobuilder_variables[auto_var]] = tower.autobuilder_values[auto_var]
                current_values = find_tower(autobuilder_results, current_tower, header_dict)
                if current_values[0] == None:
                    current_values[0]
                else:
                    progression.tower_numbers.append(current_values[0])
                    progression.period.append(current_values[1])
                    progression.cr.append(current_values[2])
                    progression.cost.append(current_values[3])
        
        
        #fill in raw scores and normalized scores to progressions
        print('recording progression scores...')
        
        a = 0 
        b = 100
            
        for progression in progression_list:
            progression.dev_period = std_deviation_period(progression)
            progression.mean_cost  = mean_cost(progression)
            progression.mean_cr = mean_cr(progression)
        
            #normal scaled of std period
            normal_period_scale = a + (progression.dev_period - normal_period_A)*(b-a)/(normal_period_B - normal_period_A)
            if normal_period_scale > 100:
                normal_period_scale = 100
            if normal_period_scale < 0:
                normal_period_scale = 0
            progression.normal_period = normal_period_scale    
                
            #normal scaled of mean cr
            normal_cr_scale = a + (progression.mean_cr - normal_cr_A)*(b-a)/(normal_cr_B - normal_cr_A)
            if normal_cr_scale > 100:
                normal_cr_scale = 100
            if normal_cr_scale < 0:
                normal_cr_scale = 0        
            progression.normal_cr = normal_cr_scale  
               
            #normal scaled of mean cost
            normal_cost_scale = a + (progression.mean_cost - normal_cost_A)*(b-a)/(normal_cost_B - normal_cost_A)
            if normal_cost_scale > 100:
                normal_cost_scale = 100
            if normal_cost_scale < 0:
                normal_cost_scale = 0            
            progression.normal_cost = normal_cost_scale 
            
        #excel output
        print('writing results to excel...')
        
        wb = Workbook()
        
        # grab the active worksheet
        ws = wb.active
        
        # Rows can also be appended
        ws.append(["Progression Number", "Variable Order", "Tower Numbers", "Periods", "CR", "Cost", "Std Dev Period", "Mean CR", "Mean Cost", "Normalized Period", "Normalized CR", "Normalized Cost"])
        
        output_count = 1
        for prog in progression_list:
            ws.append([output_count , ', '.join(map(str,prog.variable_order)), ', '.join(map(str, prog.tower_numbers)), ', '.join(map(str,prog.period)),', '.join(map(str,prog.cr)),', '.join(map(str,prog.cost)), prog.dev_period, prog.mean_cr, prog.mean_cost, prog.normal_period, prog.normal_cr, prog.normal_cost])
            output_count += 1
            
        
        # Save the file
        wb.save(output_sheet)    
        print('RUN COMPLETE')
        
runTowerProgressions = TowerProgressions()